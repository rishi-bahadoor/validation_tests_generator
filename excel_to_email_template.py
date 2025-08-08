#!/usr/bin/env python3

import sys
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from email.message import EmailMessage
from email import policy

#==============================================================================
# Configuration
#==============================================================================

DEFAULT_XLSX    = "test_report.xlsx"
EML_OUTPUT      = "test_report_email.eml"

EMAIL_FROM_DEFAULT     = "noreply@example.com"
EMAIL_TO_DEFAULT       = "qa-team@example.com"
EMAIL_CC        = ""    # comma-separated; empty = no CC

EMAIL_SUBJECT   = "Validation Report"

OPENING_LINES = [
    "Hello team,",
    "Please find the latest validation report below. Let me know if you have any questions."
]

# Your exact data headers (used to detect the main table)
HEADERS = [
    "Test ID",
    "Test Group",
    "Priority",
    "Description",
    "Pass Condition",
    "Status",
    "Notes"
]

# Re-use your Excel conditional-formatting colours
STATUS_COLORS = {
    "Pass":    "00C851",   # Green
    "Fail":    "ff4444",   # Red
    "Pending": "ffbb33",   # Orange
    "Blocked": "33b5e5",   # Blue
}

# Sheets that don’t have the above headers and should be rendered generically
GENERIC_SHEETS = {"Technician_Issues", "Github_Issues"}

#==============================================================================
# Helpers
#==============================================================================

def normalize(val):
    return val.strip().lower() if isinstance(val, str) else ""

def find_header_row(ws):
    """Return the row index (1-based) where all HEADERS appear, or None."""
    for r in range(1, min(ws.max_row, 50) + 1):
        row_vals = [normalize(ws.cell(r, c).value) 
                    for c in range(1, ws.max_column + 1)]
        if all(normalize(h) in row_vals for h in HEADERS):
            return r
    return None

def row_is_blank(row):
    return all(cell.value in (None, "") for cell in row)

def style_attrs(cell):
    """
    Inline CSS for:
     - bold font
     - wrap_text
     - Status fills from STATUS_COLORS
    """
    css = []

    if cell.font and cell.font.bold:
        css.append("font-weight:bold;")
    if cell.alignment and cell.alignment.wrap_text:
        css.append("white-space:pre-wrap;")

    # Apply color based on the cell.value → STATUS_COLORS mapping
    v = cell.value
    if isinstance(v, str) and v in STATUS_COLORS:
        css.append(f"background-color:#{STATUS_COLORS[v]};")

    return f' style="{"".join(css)}"' if css else ""

def column_widths(ws, n_cols):
    """
    Return a list of pixel widths for the first n_cols columns,
    approximating Excel's width * 7 + 5px padding.
    """
    out = []
    for c in range(1, n_cols + 1):
        letter = get_column_letter(c)
        w = ws.column_dimensions[letter].width or 8.43
        px = int(w * 7 + 5)
        out.append(px)
    return out

def render_data_sheet(ws):
    """
    Render sheets that contain one or more “header blocks”:
      GROUP_LABEL  ← one cell above
      header row   ← matches HEADERS
      data rows    ← until a blank row
    Returns a single HTML snippet for that sheet, or None.
    """
    snippets = []
    max_row = ws.max_row
    n_cols  = len(HEADERS)  # assume real data cols >= your HEADERS count

    r = 1
    while r <= max_row:
        # 1) find next header row at or after r
        hdr = None
        for rr in range(r, min(max_row, r + 100)):  # lookahead window
            vals = [normalize(ws.cell(rr, c).value)
                    for c in range(1, ws.max_column + 1)]
            if all(normalize(h) in vals for h in HEADERS):
                hdr = rr
                break
        if hdr is None:
            break

        # 2) grab the group label above
        group_label = ws.cell(row=hdr - 1, column=1).value or ""

        # 3) collect data rows (header + below) until a blank row
        block = []
        for rr in range(hdr, max_row + 1):
            row = list(ws[rr])
            if row_is_blank(row):
                break
            block.append(row)

        if not block:
            r = rr + 1
            continue

        # 4) build HTML for this block
        widths = column_widths(ws, len(block[0]))
        html = []
        html.append('<table border="1" cellpadding="4" cellspacing="0">')
        html.append("<colgroup>")
        for w in widths:
            html.append(f'<col style="width:{w}px;"/>')
        html.append("</colgroup>")

        # the GROUP label as a full-width bold row
        if group_label:
            html.append("<thead>")
            html.append(
                f'<tr>'
                f'<th colspan="{len(block[0])}" '
                f'style="font-weight:bold; text-align:left;">'
                f'{group_label}'
                f'</th>'
                f'</tr>'
            )
        else:
            html.append("<thead>")

        # the actual header row
        html.append("<tr>")
        for cell in block[0]:
            text = cell.value or ""
            html.append(f'<th{style_attrs(cell)}>{text}</th>')
        html.append("</tr>")
        html.append("</thead>")

        # body rows
        html.append("<tbody>")
        for row in block[1:]:
            html.append("<tr>")
            for cell in row:
                text = cell.value or ""
                html.append(f'<td{style_attrs(cell)}>{text}</td>')
            html.append("</tr>")
        html.append("</tbody>")
        html.append("</table><br/>")

        snippets.append("".join(html))

        # 5) advance r past this block
        r = rr + 1

    return "".join(snippets) if snippets else None

def render_generic_sheet(ws):
    """
    Render sheets like Technician_Issues or Github_Issues 
    as a simple table of every non-blank row.
    """
    rows = [row for row in ws.iter_rows() if not row_is_blank(row)]
    if not rows:
        return None

    html = ['<table border="0" cellpadding="4" cellspacing="0">']
    for row in rows:
        html.append("<tr>")
        for cell in row:
            if cell.value is not None:
                text = cell.value
                html.append(f'<td{style_attrs(cell)}>{text}</td>')
        html.append("</tr>")
    html.append("</table><br/>")

    return "".join(html)

def extract_tables(path):
    wb = load_workbook(path, data_only=False)
    result = []

    for ws in wb.worksheets:
        # 1) Main data sheets
        snippet = render_data_sheet(ws)
        if snippet:
            result.append(snippet)
            continue

        # 2) Generic sheets
        if ws.title in GENERIC_SHEETS:
            snippet = render_generic_sheet(ws)
            if snippet:
                result.append(snippet)

    return result

#==============================================================================
# Email assembly
#==============================================================================

def build_body(tables):
    html = ["<html><body>"]

    # === your opening statement ===
    for line in OPENING_LINES:
        html.append(f"<p>{line}</p>")
    html.append("<br/>")
    # ===============================

    html.extend(tables)
    html.append("</body></html>")
    return "".join(html)

def make_email(html, from_addr, to_addr):
    msg = EmailMessage(policy=policy.SMTP)
    msg["Subject"] = EMAIL_SUBJECT
    msg["From"]    = from_addr
    msg["To"]      = to_addr
    if EMAIL_CC.strip():
        msg["Cc"] = EMAIL_CC

    msg.set_content("This report is best viewed in an HTML email client.")
    msg.add_alternative(html, subtype="html")
    return msg

def save_eml(msg, path):
    with open(path, "wb") as f:
        f.write(msg.as_bytes())

#==============================================================================
# Main
#==============================================================================

def main():
    # if two args passed → override defaults
    if len(sys.argv) == 3:
        from_addr = sys.argv[1]
        to_addr   = sys.argv[2]
    elif len(sys.argv) == 1:
        from_addr = EMAIL_FROM_DEFAULT
        to_addr   = EMAIL_TO_DEFAULT
    else:
        print(f"Usage: {sys.argv[0]} [sender_email recipient_email]")
        sys.exit(1)

    if not os.path.exists(DEFAULT_XLSX):
        print(f"Error: `{DEFAULT_XLSX}` not found.")
        sys.exit(1)

    tables = extract_tables(DEFAULT_XLSX)
    body   = build_body(tables)
    msg    = make_email(body, from_addr, to_addr)
    save_eml(msg, EML_OUTPUT)
    print(f"✅ Generated `{EML_OUTPUT}`. Verify in Outlook.")


if __name__ == "__main__":
    main()