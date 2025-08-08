import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule

## Constants ##
HEADERS = [
    ("Test ID",       22),
    ("Test Group",    18),
    ("Priority",      10),
    ("Description",   90),
    ("Pass Condition",30),
    ("Status",        10),
    ("Notes",         50),
]

STATUS_OPTIONS = ["Pass", "Fail", "Pending", "Blocked"]
STATUS_COLORS  = {
    "Pass":    "00C851",
    "Fail":    "ff4444",
    "Pending": "ffbb33",
    "Blocked": "33b5e5",
}

## Helpers ##
def find_header_rows(ws, header_names):
    """
    Return a sorted list of all row indices where the first
    len(header_names) cells match header_names (case-insensitive).
    """
    target = [h.strip().lower() for h in header_names]
    hits = []
    for row in range(1, ws.max_row + 1):
        cells = [
            (ws.cell(row=row, column=col).value or "").strip().lower()
            for col in range(1, len(header_names) + 1)
        ]
        if cells == target:
            hits.append(row)
    return hits

def find_column_letter_at_row(ws, header_name, header_row):
    """
    In a specific header_row, find the column letter whose cell equals header_name.
    """
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and str(val).strip().lower() == header_name.strip().lower():
            return get_column_letter(col)
    raise ValueError(f"Header '{header_name}' not found in row {header_row}")

def wrap_all_cells(ws):
    wrap_alignment = Alignment(wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap_alignment

def add_sheet(wb, name):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(title=name)

## Main formatting ##
def format_excel_sheet(path):
    wb = load_workbook(path)
    ws = wb.active

    # 1) Discover every header row in the sheet
    header_names = [h for h, _ in HEADERS]
    header_rows  = find_header_rows(ws, header_names)
    if not header_rows:
        raise RuntimeError("No header row matching HEADERS found!")

    # 2) Apply column-widths globally (same columns each region)
    for header, width in HEADERS:
        letter = find_column_letter_at_row(ws, header, header_rows[0])
        ws.column_dimensions[letter].width = width

    # 3) For each table region, add dropdown + conditional formatting + bold headers
    #    The data region is from header_row+1 up to just before the next header_row (or sheet end)
    sentinel = header_rows + [ws.max_row + 1]
    for idx, header_row in enumerate(header_rows):
        start_row = header_row + 1
        end_row   = sentinel[idx + 1] - 1

        # 3a) Bold the header row
        for col in range(1, ws.max_column + 1):
            ws.cell(row=header_row, column=col).font = Font(bold=True)

        # 3b) Dropdown in “Status” column for this region
        status_col = find_column_letter_at_row(ws, "Status", header_row)
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(STATUS_OPTIONS)}"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        for r in range(start_row, end_row + 1):
            dv.add(f"{status_col}{r}")

        # 3c) Conditional formatting for the same region
        for value, hex_color in STATUS_COLORS.items():
            fill = PatternFill(
                start_color=hex_color,
                end_color=hex_color,
                fill_type="solid"
            )
            # Check cell in first data row == value
            formula = [f'${status_col}{start_row}="{value}"']
            cell_range = f"{status_col}{start_row}:{status_col}{end_row}"
            rule = FormulaRule(formula=formula, fill=fill)
            ws.conditional_formatting.add(cell_range, rule)

    # 4) Wrap all cells, add your extra sheets
    wrap_all_cells(ws)

    tech_ws   = add_sheet(wb, "Technician_Issues")
    populate_single_cell(tech_ws, "Technician Issues")
    github_ws = add_sheet(wb, "Github_Issues")
    populate_single_cell(github_ws, "Github Issues")

    wb.save(path)
    print(f"✅ Excel formatting applied to: {path}")

## Small helper kept in place ##
def populate_single_cell(ws, title_text):
    ws["A1"] = title_text
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 120
    wrap_all_cells(ws)


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python excel_format.py <path_to_excel_file>")
        sys.exit(1)
    format_excel_sheet(sys.argv[1])