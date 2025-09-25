# VERSION 1.1.5

import sys
import win32com.client
import time

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule

from pathlib import Path


## Constants ##
HEADERS = [
    ("Test ID",       22),
    ("Test Group",    18),
    ("Priority",      10),
    ("Description",   90),
    ("Status",        10),
    ("Notes",         50),
    ("Frequency", 20),
]

STATUS_OPTIONS = ["Pass", "Fail", "Pending", "Blocked"]
STATUS_COLORS  = {
    "Pass":    "c6efce",
    "Fail":    "ffc7ce",
    "Pending": "ffeb9c",
    "Blocked": "add8e6",
}

## Helpers ##
def force_close_excel_file(path):
    excel = win32com.client.Dispatch("Excel.Application")
    for wb in excel.Workbooks:
        if wb.FullName.lower() == str(path).lower():
            wb.Close(SaveChanges=True)
            break
    excel.Quit()
    time.sleep(4)

def reopen_excel_file(path):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True
    excel.WindowState = -4137
    excel.Workbooks.Open(str(path))

def find_header_rows(ws, header_names):
    """
    Return a sorted list of all row indices where the first
    len(header_names) cells match header_names (case-insensitive).
    """
    target = [h.strip().lower() for h in header_names]
    hits = []
    for row in range(1, ws.max_row + 1):
        cells = [
            (ws.cell(row=row, column=col).value or "").strip().lower()
            for col in range(1, len(header_names) + 1)
        ]
        if cells == target:
            hits.append(row)
    return hits

def find_column_letter_at_row(ws, header_name, header_row):
    """
    In a specific header_row, find the column letter whose cell equals header_name.
    """
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and str(val).strip().lower() == header_name.strip().lower():
            return get_column_letter(col)
    raise ValueError(f"Header '{header_name}' not found in row {header_row}")

def wrap_all_cells(ws):
    wrap_alignment = Alignment(wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap_alignment

def add_sheet(wb, name):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(title=name)

## Main formatting ##
def format_excel_sheet(path):
    wb = load_workbook(path)
    ws = wb.active

    # 1) Discover every header row in the sheet
    header_names = [h for h, _ in HEADERS]
    header_rows  = find_header_rows(ws, header_names)
    if not header_rows:
        raise RuntimeError("No header row matching HEADERS found!")

    # 2) Apply column-widths globally (same columns each region)
    for header, width in HEADERS:
        letter = find_column_letter_at_row(ws, header, header_rows[0])
        ws.column_dimensions[letter].width = width

    # 3) For each table region, add dropdown + conditional formatting + bold headers
    #    The data region is from header_row+1 up to just before the next header_row (or sheet end)
    sentinel = header_rows + [ws.max_row + 1]
    for idx, header_row in enumerate(header_rows):
        start_row = header_row + 1
        end_row   = sentinel[idx + 1] - 1

        # 3a) Bold the header row
        for col in range(1, ws.max_column + 1):
            ws.cell(row=header_row, column=col).font = Font(bold=True)

        # 3b) Dropdown in “Status” column for this region
        status_col = find_column_letter_at_row(ws, "Status", header_row)
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(STATUS_OPTIONS)}"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        for r in range(start_row, end_row + 1):
            dv.add(f"{status_col}{r}")

        # 3c) Conditional formatting for the same region
        for value, hex_color in STATUS_COLORS.items():
            fill = PatternFill(
                start_color=hex_color,
                end_color=hex_color,
                fill_type="solid"
            )
            # Check cell in first data row == value
            formula = [f'${status_col}{start_row}="{value}"']
            cell_range = f"{status_col}{start_row}:{status_col}{end_row}"
            rule = FormulaRule(formula=formula, fill=fill)
            ws.conditional_formatting.add(cell_range, rule)

    # 4) Wrap all cells, add your extra sheets
    wrap_all_cells(ws)

    tech_ws   = add_sheet(wb, "Technician_Issues")
    populate_single_cell(tech_ws, "Technician Issues")

    wb.save(path)

## Small helper kept in place ##
def populate_single_cell(ws, title_text):
    ws["A1"] = title_text
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 120
    wrap_all_cells(ws)

def update_test_row_by_id(ws, test_id: str, new_status: str, new_notes: str = None):
    """
    Update the Status and optionally Notes cell for the row matching the given Test ID.
    """
    if new_status not in STATUS_OPTIONS:
        raise ValueError(f"Invalid status '{new_status}'. Must be one of: {STATUS_OPTIONS}")

    header_names = [h for h, _ in HEADERS]
    header_rows = find_header_rows(ws, header_names)
    if not header_rows:
        raise RuntimeError("No header row matching HEADERS found!")

    test_id_col = find_column_letter_at_row(ws, "Test ID", header_rows[0])
    status_col  = find_column_letter_at_row(ws, "Status", header_rows[0])
    notes_col   = find_column_letter_at_row(ws, "Notes", header_rows[0])

    for row in range(header_rows[0] + 1, ws.max_row + 1):
        cell_value = str(ws[f"{test_id_col}{row}"].value).strip()
        if cell_value == test_id:
            ws[f"{status_col}{row}"].value = new_status
            if new_notes is not None:
                ws[f"{notes_col}{row}"].value = new_notes
            return

    raise ValueError(f"Test ID '{test_id}' not found in sheet.")

def safe_update_excel(path, test_id, new_status, new_notes=None):
    path = Path(path).resolve()

    try:
        force_close_excel_file(path)
    except Exception as e:
        print(f"⚠️ Could not close Excel file: {e}")

    wb = load_workbook(path)
    ws = wb.active
    update_test_row_by_id(ws, test_id, new_status, new_notes)
    wb.save(path)

    try:
        reopen_excel_file(path)
    except Exception as e:
        print(f"⚠️ Could not reopen Excel file: {e}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage:")
        print("  Format: python excel_format.py format <path_to_excel_file>")
        print("  Update: python excel_format.py update <path_to_excel_file> <test_id> <new_status> [notes]")
        sys.exit(1)

    mode = sys.argv[1].lower()
    path = sys.argv[2]

    if mode == "format":
        format_excel_sheet(path)

    elif mode == "update":
        if len(sys.argv) < 5:
            print("Usage: python excel_format.py update <path_to_excel_file> <test_id> <new_status> [notes]")
            sys.exit(1)
        test_id = sys.argv[3]
        new_status = sys.argv[4]
        new_notes = sys.argv[5] if len(sys.argv) >= 6 else None
        safe_update_excel(path, test_id, new_status, new_notes)

    else:
        print(f"Unknown mode '{mode}'. Use 'format' or 'update'.")
        sys.exit(1)