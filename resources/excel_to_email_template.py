# VERSION 1.1.2

import sys
from pathlib import Path

from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from datetime import date

# =============================================================================
# Configuration Defaults
# =============================================================================
EMAIL_FROM_DEFAULT    = "no-reply@example.com"
EMAIL_TO_DEFAULT      = ["alice@example.com", "bob@example.com"]
EXCEL_FILE_DEFAULT    = "validation_test_report.xlsx"
IMAGES_FOLDER_DEFAULT = "images_github_issues/"
IMAGES_FOLDER_MISC    = "images_misc/"

STATUS_COLORS = {
    "Pass":    "#c6efce",
    "Fail":    "#ffc7ce",
    "Pending": "#ffeb9c",
    "Blocked": "#add8e6",
}

EMAIL_FILE_NAME = "validation_report_message.eml"
EMAIL_SUBJECT   = f"UFB Ultra Release Status Track - {date.today().strftime('%d-%m-%Y')}"

# =============================================================================
# Small Helpers
# =============================================================================

def is_empty_row(row):
    return all(cell.value in (None, "") for cell in row)

def get_blocks_from_sheet(sheet):
    """
    Split a worksheet into blocks separated by empty rows.
    Each block is a list of openpyxl row tuples.
    """
    blocks, current = [], []
    for row in sheet.iter_rows():
        if is_empty_row(row):
            if current:
                blocks.append(current)
                current = []
        else:
            current.append(row)
    if current:
        blocks.append(current)
    return blocks

def is_label_row(row):
    """
    Detect a subtable label row: exactly one non-empty cell in col A.
    """
    first, *rest = row
    return (
        first.value not in (None, "")
        and all(cell.value in (None, "") for cell in rest)
    )

def split_subtables(block):
    """
    Given a list of rows (no empty rows), split into subtables:
      [label, header, data..., label2, header2, data...]
    Returns a list of { label: str, rows: [ {hdr: val}, ... ] }.
    """
    subtables, i, n = [], 0, len(block)
    while i < n:
        if is_label_row(block[i]):
            label = block[i][0].value
            i += 1
            headers = [c.value for c in block[i]]
            i += 1
            data_rows = []
            while i < n and not is_label_row(block[i]):
                data_rows.append(block[i])
                i += 1
            rows = [
                {headers[j]: data_row[j].value for j in range(len(headers))}
                for data_row in data_rows
            ]
            subtables.append({"label": label, "rows": rows})
        else:
            i += 1
    return subtables

def extract_table_data(xlsx_path, sheet_index, block_index, data_type):
    """
    Grab blocks from the given sheet.
    - data_type "meta"     : return single block as List[(key, val)]
    - data_type "table"    : return one block as List[Dict[hdr, val]]
    - data_type "subtables": return ALL blocks from block_index onward,
                             each as a labeled subtable dict.
    """
    wb    = load_workbook(xlsx_path, data_only=True)
    sheet = wb.worksheets[sheet_index - 1]
    blocks = get_blocks_from_sheet(sheet)

    if data_type == "meta":
        block = blocks[block_index - 1]
        return [(r[0].value, r[1].value) for r in block]

    if data_type == "subtables":
        subtables = []
        # every block from block_index onward is one (or more) subtable(s)
        for block in blocks[block_index - 1 :]:
            subtables.extend(split_subtables(block))
        return subtables

    # data_type == "table" (legacy single-table)
    block = blocks[block_index - 1]
    headers = [c.value for c in block[0]]
    return [
        {headers[j]: row[j].value for j in range(len(headers))}
        for row in block[1:]
    ]

def apply_color_to_cells(rows, column_name, color_map):
    """
    Annotate each row with a color for the given column.
    Adds "__color_<column_name>" to each row dict when a color matches.
    """
    colored = []
    for row in rows:
        new_row = row.copy()
        val     = row.get(column_name)
        color   = color_map.get(val)
        if color:
            new_row[f"__color_{column_name}"] = color
        colored.append(new_row)
    return colored


# =============================================================================
# Functional Code: EmailBodyBuilder
# =============================================================================

class EmailBodyBuilder:
    def __init__(self, subject, sender, recipients):
        self.msg           = MIMEMultipart("related")
        self.msg["Subject"] = subject
        self.msg["From"]    = sender
        self.msg["To"]      = ", ".join(recipients)
        self._html_parts   = []

    def email_body_add_intro(self, lines):
        html = ""
        for line in lines:
            html += f"<p style='font-family:Arial; margin:4px 0'>{line}</p>"
        self._html_parts.append(html)

    def email_body_start(self, title: str):
        self._html_parts.append(f"<h2 style='font-family:Arial'>{title}</h2><hr>")

    def email_body_add_table(self, data, data_type="table"):
        """
        Render:
          - meta     : List[(k,v)]
          - table    : List[Dict]
          - subtables: List[{label, rows}]
        """
        if data_type == "meta":
            html = "<table style='font-family:Arial; margin-bottom:20px;'>"
            for k,v in data:
                html += (
                    "<tr>"
                    f"<td style='font-weight:bold; padding:4px'>{k}</td>"
                    f"<td style='padding:4px'>{v}</td>"
                    "</tr>"
                )
            html += "</table>"

        elif data_type == "subtables":
            html = ""
            for sub in data:
                html += f"<h4 style='font-family:Arial;margin:8px 0 4px'>{sub['label']}</h4>"
                html += self._render_plain_table(sub["rows"])
            html = f"<div style='margin-bottom:20px'>{html}</div>"

        else:  # plain table
            html = self._render_plain_table(data)

        self._html_parts.append(html)

    def _render_plain_table(self, rows):
        if not rows:
            return "<p><em>No data</em></p>"

        headers = [
            h for h in rows[0].keys()
            if isinstance(h, str) and not h.startswith("__color_")
        ]
        html = (
            "<table border='1' cellpadding='4' "
            "style='border-collapse:collapse;font-family:Arial;width:100%;'>"
        )
        html += "<tr>" + "".join(f"<th style='background:#efefef'>{h}</th>" for h in headers) + "</tr>"

        for row in rows:
            html += "<tr>"
            for h in headers:
                raw   = row.get(h)
                val   = "" if raw is None else raw
                color = row.get(f"__color_{h}")
                style = f"background:{color};" if color else ""
                html += f"<td style='padding:4px;{style}'>{val}</td>"
            html += "</tr>"

        html += "</table>"
        return html

    def email_body_inline_images_from_folder(self):
        """
        Embed images from the default folder, inlined by CID.
        """
        for img_path in Path(IMAGES_FOLDER_DEFAULT).glob("*.*"):
            cid = img_path.stem
            img = MIMEImage(img_path.read_bytes(), _subtype=img_path.suffix.lstrip("."))
            img.add_header("Content-ID", f"<{cid}>")
            self.msg.attach(img)
            self._html_parts.append(
                f"<div><img src='cid:{cid}' style='max-width:600px;'/></div>"
            )

    def email_body_attach_images(self, folder):
        """
        Attach images from a folder as file attachments (not inline).
        """
        for img_path in Path(folder).glob("*.*"):
            img = MIMEImage(img_path.read_bytes(), _subtype=img_path.suffix.lstrip("."))
            img.add_header("Content-Disposition", "attachment", filename=img_path.name)
            self.msg.attach(img)

    def email_body_close(self):
        full_html = "<html><body>" + "".join(self._html_parts) + "</body></html>"
        self.msg.attach(MIMEText(full_html, "html"))
        return self.msg


# =============================================================================
# Core Email Builder
# =============================================================================
def build_email_message(xlsx_path, from_addr, recipients):
    # extract metadata (block 1)
    meta_data = extract_table_data(
        xlsx_path, sheet_index=1, block_index=1, data_type="meta"
    )

    # extract ALL subtable-blocks from sheet1, starting at block 2
    main_subtables = extract_table_data(
        xlsx_path, sheet_index=1, block_index=2, data_type="subtables"
    )

    # extract a plain table from sheet2 (single block)
    secondary_tbl = extract_table_data(
        xlsx_path, sheet_index=2, block_index=1, data_type="table"
    )

    # apply coloring to each subtable's Status column
    for sub in main_subtables:
        sub["rows"] = apply_color_to_cells(sub["rows"], "Status", STATUS_COLORS)

    builder = EmailBodyBuilder(
        subject=EMAIL_SUBJECT,
        sender=from_addr,
        recipients=recipients
    )
    builder.email_body_add_intro([
        "Hello team,",
        "Please find the latest validation report below. Let me know if you have any questions."
    ])

    builder.email_body_start("")
    builder.email_body_add_table(meta_data, data_type="meta")

    # inline images from default folder
    builder.email_body_inline_images_from_folder()

    # Attach misc images as file attachments
    builder.email_body_attach_images(IMAGES_FOLDER_MISC)

    # Render subtables and secondary table
    builder.email_body_add_table(main_subtables, data_type="subtables")
    builder.email_body_add_table(secondary_tbl, data_type="table")

    return builder.email_body_close()


# =============================================================================
# Main Entry Point
# =============================================================================

def main():
    if len(sys.argv) == 4:
        from_addr   = sys.argv[1]
        to_addr     = sys.argv[2]
        excel_file  = sys.argv[3]
    elif len(sys.argv) == 1:
        from_addr = EMAIL_FROM_DEFAULT
        to_addr   = EMAIL_TO_DEFAULT
        excel_file = EXCEL_FILE_DEFAULT
    else:
        print(f"Usage: {sys.argv[0]} [sender_email recipient_email]")
        sys.exit(1)

    recipients = to_addr if isinstance(to_addr, list) else [to_addr]
    message    = build_email_message(excel_file, from_addr, recipients)

    with open(EMAIL_FILE_NAME, "w") as f:
        f.write(message.as_string())


if __name__ == "__main__":
    main()